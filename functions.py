# -*- coding: utf-8 -*-
"""
Created on Wed Dec  2 16:04:51 2020

@author: Anselmettic
"""
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_interval, get_column_letter
import re
import openpyxl
import os
import datetime as dt
from datetime import date, timedelta
import datetime
abspath = os.path.abspath(__file__)
dname = os.path.dirname(abspath)
os.chdir(dname)
#
def get_config():
    config = pd.read_csv("config.csv",sep=",",header=None)
    config = config.set_index(0)
    config = config.drop(2,axis=1)
    return config
        
def get_energydemand(ws,years,scenarios,applications,ener_sources,aggregates):
    applications = applications + aggregates
### Sort to extract correct locations later    
    sort = {}
    for row in ws.iter_rows():
        for app in applications:
            for cell in row:
                if cell.value == app:
                    sort[app] = cell.row
    applications = sorted(sort,key=sort.get)
### Both scenarios displayed next to each other in same spreadsheet. Get location of values for each scenario. ###
    locs_s  = {}
    for col in ws.iter_cols(min_row=1, max_row=50, max_col=20):
            for cell in col:
                for scen in scenarios:
                    if cell.value==scen:
                        locs_s[scen+'_row'] = int(cell.row)
                        locs_s[scen+'_col'] = int(cell.col_idx)
    diff = locs_s[scenarios[1]+'_col'] - locs_s[scenarios[0]+'_col']
### Get location of values for each application. ###    
    locs = {scen: {app: pd.DataFrame(index=['start','end'],columns=['col','row']) for app in applications} for scen in scenarios}
    index = 0
    ref = ''  
    for scen in scenarios: 
        for row in ws.iter_rows(max_row=200,
                                min_col=locs_s[scen+'_col'],
                                max_col=locs_s[scen+'_col']+diff-1):
            for app in applications:
                    for cell in row:
                        if cell.value==app:
                            locs[scen][app].loc['start','col'] = cell.col_idx
                            locs[scen][app].loc['start','row'] = cell.row
                    index = applications.index(app)
                    locs[scen][app].loc['end','col'] = locs[scen][app].loc['start','col'] + diff
                    if index < len(applications)-1:
                        ref = applications[index+1]
                        locs[scen][app].loc['end','row'] = locs[scen][ref].loc['start','row'] - 1
                    else:
                        locs[scen][app].loc['end','row'] = locs[scen][ref].loc['start']['row'] + 200
### else-Schleife hier ist fehleranfällig. Es ist wichtig, dass sämtliche 'applications' aufgeführt sind, sodass Datenbereich für letzte Application nicht noch andere Daten miteinbezieht.
### Evtl. allgemeingültigere Lösung finden. 
    ### GET DATA
    df = {scen: {app: pd.DataFrame(index=ener_sources,columns=years) for app in applications} for scen in scenarios}
    row = 0
    col_w = 0
    for scen in scenarios:
        for app in applications:
            for en in ener_sources:
                for row in ws.iter_rows(min_row=locs[scen][app].loc['start','row'],
                                     max_row=locs[scen][app].loc['end','row'],
                                     min_col=locs[scen][app].loc['start','col'],
                                     max_col=locs[scen][app].loc['end','col']):
                    for cell in row:
                        if cell.value==en:
                            row = int(cell.row)
                            for yr in years:
                                for col in ws.iter_cols(min_row=locs[scen][app].loc['start','row'],
                                                 max_row=locs[scen][app].loc['end','row'],
                                                 min_col=locs[scen][app].loc['start','col'],
                                                 max_col=locs[scen][app].loc['end','col']):
                                    for cell_c in col:
                                        if cell_c.value == yr:
                                            col_w = int(cell_c.col_idx)
                                            if ws.cell(row,col_w).value is None:
                                                df[scen][app].loc[en,yr] = 0.0
                                            else:
                                                df[scen][app].loc[en,yr] = float(ws.cell(row,col_w).value)
    return df

def aggregate_demand(df,years,scenarios,applications,ener_sources):
#    applications = applications + aggregates
#    # drop aggregates
#    for agg in aggregates:
#        for scen in scenarios:
#            del(df[scen][agg])
#        index = applications.index(agg)
#        del(applications[index])
#   deactivated this part since it was prone to producing errors. it was written, when there was no distinction between aggregates + applications and were both saved in same dataframe, thus code was laborious. this should facilitate things.     
    ## aggregate demand
    agg_energy_demand = {scen: pd.DataFrame(index=ener_sources,columns=years) for scen in scenarios}
    for scen in scenarios:
        for app in applications:
                agg_energy_demand[scen] = agg_energy_demand[scen].add(df[scen][app],fill_value=0.0)
    return agg_energy_demand

def intrapolate(full_energydemand,scenarios,index_agg,sector):
    for scen in scenarios:
        for idx in index_agg:
            full_energydemand[scen].loc[idx,2055] = full_energydemand[scen].loc[idx,2050]
            full_energydemand[scen].loc[idx,2060] = full_energydemand[scen].loc[idx,2050]
            full_energydemand[scen].loc[idx,2065] = full_energydemand[scen].loc[idx,2050]
            full_energydemand[scen].loc[idx,2070] = full_energydemand[scen].loc[idx,2050]
            if sector != 'buildings' and sector != 'heat':
                if full_energydemand[scen].loc[idx,2018] != 0:
                    full_energydemand[scen].loc[idx,2019] = ((full_energydemand[scen].loc[idx,2025]/full_energydemand[scen].loc[idx,2018])**(1/7))**1 *  full_energydemand[scen].loc[idx,2018]
                    full_energydemand[scen].loc[idx,2020] = ((full_energydemand[scen].loc[idx,2025]/full_energydemand[scen].loc[idx,2018])**(1/7))**2 *  full_energydemand[scen].loc[idx,2018]
                else:
                    full_energydemand[scen].loc[idx,2019] = 0
                    full_energydemand[scen].loc[idx,2020] = 0
                if full_energydemand[scen].loc[idx,2030] != 0:
                    full_energydemand[scen].loc[idx,2035] = ((full_energydemand[scen].loc[idx,2040]/full_energydemand[scen].loc[idx,2030])**(1/10))**5 *  full_energydemand[scen].loc[idx,2030]
                else:
                    full_energydemand[scen].loc[idx,2035] = 0
                if full_energydemand[scen].loc[idx,2040] != 0:
                    full_energydemand[scen].loc[idx,2045] = ((full_energydemand[scen].loc[idx,2050]/full_energydemand[scen].loc[idx,2040])**(1/10))**5 *  full_energydemand[scen].loc[idx,2040]
                else:
                    full_energydemand[scen].loc[idx,2045] = 0
            else:
                full_energydemand[scen].loc[idx,2018] = full_energydemand[scen].loc[idx,2019]
    return full_energydemand
        
def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

def get_energydemand_industry(sheet,file,extraction_years,scenarios,applications,ener_sources,aggregates,dname):
    applications = applications + aggregates
    locs = {}
    locs_frame = {scen: {app: pd.DataFrame(0, index=['start','end'],columns=['col','row']) for app in applications} for scen in scenarios}
    df = {scen: {app: pd.DataFrame(index=ener_sources,columns=extraction_years) for app in applications} for scen in scenarios}   
    row_w = 0
    col_w = 0
    for scen in scenarios:
        wb = openpyxl.load_workbook(dname+file[scen], data_only=True)
        ws = wb[sheet]
        ### SORT to avoid miscalculating the locations. 
        sort = {}
        for row in ws.iter_rows():
            for app in applications:
                for cell in row:
                    if cell.value == app:
                        sort[app] = cell.row
                        locs[app + '_' + scen] = int(cell.row)
        applications = sorted(sort,key=sort.get)     
        ## Grab matrices 
        index = 0
        ref = ''  
        for row in ws.iter_rows():
            for app in applications:
                for cell in row:
                    if cell.value==app:
                        locs_frame[scen][app].loc['start','col'] = cell.col_idx
                        locs_frame[scen][app].loc['start','row'] = cell.row
                    if cell.value==2070:
                        locs_frame[scen][app].loc['end','col'] = cell.col_idx
                index = applications.index(app)
                if index < len(applications)-1:
                    ref = applications[index+1]
                    locs_frame[scen][app].loc['end','row'] = locs_frame[scen][ref].loc['start','row'] - 1
                else:
                    locs_frame[scen][app].loc['end','row'] = locs_frame[scen][ref].loc['start']['row'] + 200    
        for app in applications:
            for en in ener_sources:
                for row in ws.iter_rows(min_row=locs_frame[scen][app].loc['start','row'],
                                 max_row=locs_frame[scen][app].loc['end','row'],
                                 min_col=locs_frame[scen][app].loc['start','col'],
                                 max_col=locs_frame[scen][app].loc['end','col']):
                    for cell in row:
                        if cell.value==en:
                            row_w = int(cell.row)
                            for yr in extraction_years:
                                for col_c in ws.iter_cols(min_row=locs_frame[scen][app].loc['start','row'],
                                                 max_row=locs_frame[scen][app].loc['end','row'],
                                                 min_col=locs_frame[scen][app].loc['start','col'],
                                                 max_col=locs_frame[scen][app].loc['end','col']):
                                    for cell_c in col_c:
                                        if cell_c.value == yr:
                                            col_w = int(cell_c.col_idx)
                                            if ws.cell(row_w,col_w).value is None:
                                                df[scen][app].loc[en,yr] = 0.0
                                            else:
                                                df[scen][app].loc[en,yr] = float(ws.cell(row_w,col_w).value)
                                        if cell_c.value == str(yr):
                                            col_w = int(cell_c.col_idx)
                                            if ws.cell(row_w,col_w).value is None:
                                                df[scen][app].loc[en,yr] = 0.0
                                            else:
                                                df[scen][app].loc[en,yr] = float(ws.cell(row_w,col_w).value)
    return df
                           
def get_profiles(sector):
    mapping = {'buildings':'elBuildings','heat':'Heat','industry':'Industry','mobility':'mobility'}
    config = get_config()
    ref_w_year = config.loc['reference weather year',1]
    ref_c_year = config.loc['reference calendar year',1]
    file = '/inputs/profiles_c' + ref_c_year + '_w' + ref_w_year + '.xlsx'
    wb = openpyxl.load_workbook(dname+file, data_only=True) # Profiles equiv. over scenarios.
    ws = wb[mapping[sector]]
    ### get location of profiles on worksheet.
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'd1':  
                for row_r in ws.iter_rows(max_row=cell.row, min_row=cell.row):
                    for cell_r in row_r:
                        if cell_r.value == 'h1':
                            start_row = cell_r.row
                            for col_c in ws.iter_cols(max_row=cell_r.row, min_row=cell_r.row):
                                for cell_c in col_c:
                                    if cell_c.value == 'd1':
                                        start_col = cell_c.col_idx
    ### BUILD RANGE.
    range_pr = get_column_letter(start_col) + str(start_row) + ":" + get_column_letter(ws.max_column) + str(ws.max_row)
    profiles = load_workbook_range(range_pr,ws)
    profiles['name'] = profiles[profiles.columns[0]] + "_" + profiles[profiles.columns[1]]
    profiles = profiles.set_index('name')
    profiles = profiles.drop(columns=[profiles.columns[0],profiles.columns[1]])
    range_colnam = get_column_letter(start_col + 2) + str(start_row - 1) + ":" + get_column_letter(ws.max_column) + str(start_row -1)
    colnam = load_workbook_range(range_colnam,ws)
    profiles.columns = colnam.stack().tolist()
    return profiles

def get_typedays():
    file = '\\inputs\\typedays.xlsx'
    wb = openpyxl.load_workbook(dname+file, data_only=True) # Profiles equiv. over scenarios.
    ws = wb['typedays']
    range_td = ""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=="d1":
                range_td = get_column_letter(cell.col_idx) + str(cell.row) + ":" + get_column_letter(cell.col_idx + 1) + str(ws.max_row)
    typedays = load_workbook_range(range_td,ws)
    return typedays

#%%%
#### WHEN2HEAT FUNCTIONS

def daily_parameters(input_path):

    file = os.path.join(input_path, 'bgw_bdew', 'daily_demand.csv')
    return pd.read_csv(file, sep=';', decimal=',', header=[0, 1], index_col=0)

def hourly_parameters(input_path):

    def read():
        file = os.path.join(input_path, 'bgw_bdew', filename)
        return pd.read_csv(file, sep=';', decimal=',', index_col=index_col).apply(pd.to_numeric, downcast='float')

    parameters = {}
    for building_type in ['SFH', 'MFH', 'COM']:

        filename = 'hourly_factors_{}.csv'.format(building_type)

        # MultiIndex for commercial heat because of weekday dependency
        index_col = [0, 1] if building_type == 'COM' else 0

        parameters[building_type] = read()

    return parameters

def upsample_df(df, resolution):

    # The low-resolution values are applied to all high-resolution values up to the next low-resolution value
    # In particular, the last low-resolution value is extended up to where the next low-resolution value would be

    df = df.copy()

    # Determine the original frequency
    freq = df.index[-1] - df.index[-2]

    # Temporally append the DataFrame by one low-resolution value
    df.loc[df.index[-1] + freq, :] = df.iloc[-1, :]

    # Up-sample
    df = df.resample(resolution).pad()

    # Drop the temporal low-resolution value
    df.drop(df.index[-1], inplace=True)

    return df

def reference_temperature(temperature):
    # Daily average
    temperature.index=pd.to_datetime(temperature.index)
    daily_average = temperature.groupby(pd.Grouper(freq='D')).mean().copy()
    # Weighted mean

    return sum([.5 ** i * daily_average.shift(i).fillna(method='bfill') for i in range(4)]) / \
           sum([.5 ** i for i in range(4)])


def daily_heat(temperature, windy_locations, all_parameters):

    # BDEW et al. 2015 describes the function for the daily heat demand
    # This is implemented in the following and passed to the general daily function

    def heat_function(t, parameters):

        #celsius = t - 273.15  # The temperature input is in Kelvin
        celsius=t

        sigmoid = parameters['A'] / (
                1 + (parameters['B'] / (celsius - 40)) ** parameters['C']
        ) + parameters['D']

        linear = pd.DataFrame(
            [parameters['m_{}'.format(i)] * celsius + parameters['b_{}'.format(i)] for i in ['s', 'w']]
        ).max()

        return sigmoid + linear

    return daily(temperature, windy_locations, all_parameters, heat_function)

def daily_water(temperature, windy_locations, all_parameters):

    # A function for the daily water heating demand is derived from BDEW et al. 2015
    # This is implemented in the following and passed to the general daily function

    def water_function(t, parameters):

        #celsius = t - 273.15  # The temperature input is in Kelvin
        celsius=t

        # Below 15 °C, the water heating demand is not defined and assumed to stay constant
        celsius.clip(15, inplace=True)

        return parameters['m_w'] * celsius + parameters['b_w'] + parameters['D']

    return daily(temperature, windy_locations, all_parameters, water_function)

def daily(temperature, windy_locations, all_parameters, func):

    # All locations are separated by the average wind speed with the threshold 4.4 m/s
#    windy_locations = {
#        'normal': wind[wind <= 4.4].index,
#        'windy': wind[wind > 4.4].index
#    }
#    
    

    buildings = ['SFH', 'MFH', 'COM']


    return pd.concat([pd.concat(
            [temperature[locations].apply(func, parameters=all_parameters[(b, windiness)])
             for windiness, locations in windy_locations.items()],
            axis=1
        ) for b in buildings],
        #keys=buildings, names=['building', 'country', 'latitude', 'longitude'], axis=1
        keys=buildings, names=['building', 'country'], axis=1
    )

def hourly_heat(daily_df, temperature, parameters):

    # According to BGW 2006, temperature classes are derived from the temperature data
    # This is re-sampled to a 60-min-resolution and passed to the general hourly function

    classes = upsample_df(
        #(np.ceil(((temperature - 273.15) / 5).astype('float64')) * 5).clip(lower=-15, upper=30),
        (np.ceil(((temperature ) / 5).astype('float64')) * 5).clip(lower=-15, upper=30),
        '60min'
    ).astype(int).astype(str)
    return hourly(daily_df, classes, parameters)

def hourly_water(daily_df, temperature, parameters):

    # For water heating, the highest temperature classes '30' is chosen
    # This is re-sampled to a 60-min-resolution and passed to the general hourly function

    classes = upsample_df(
        pd.DataFrame(30, index=temperature.index, columns=temperature.columns),
        '60min'
    ).astype(int).astype(str)

    return hourly(daily_df, classes, parameters)

def hourly(daily_df, classes, parameters):

    def hourly_factors(building):

        # This function selects hourly factors from BGW 2006 by time and temperature class
        slp = pd.DataFrame(index=classes.index, columns=classes.columns)

        # Time includes the hour of the day
        times = classes.index.map(lambda x: x.strftime('%H:%M'))
        # For commercial buildings, time additionally includes the weekday
        if building == 'COM':
            weekdays = classes.index.map(lambda x: int(x.strftime('%w')))
            times = list(zip(weekdays, times))

        for column in classes.columns:
            slp[column] = parameters[building].lookup(times, classes.loc[:, column])

        return slp

    buildings = daily_df.columns.get_level_values('building').unique()

    results = pd.concat(
        [upsample_df(daily_df, '60min')[building] * hourly_factors(building) for building in buildings],
        #keys=buildings, names=['building', 'country', 'latitude', 'longitude'], axis=1
        keys=buildings, names=['building', 'country'], axis=1
    )

    return results.swaplevel('building', 'country', axis=1)

def ewmsmoothing(ts, periods):
  return ts.ewm(span=periods).mean()

def simplesmoothing(ts, periods):
  result=[]
  help=0
  for i in range(int(len(ts)/periods)):
    avg=ts[help:help+periods].mean()
    for j in range(periods): 
      result.append(avg)
    help+=periods
  return result

#%% 
def calc_easter(year):
    "Returns Easter as a date object."
    a = year % 19
    b = year // 100
    c = year % 100
    d = (19 * a + b - b // 4 - ((b - (b + 8) // 25 + 1) // 3) + 15) % 30
    e = (32 + 2 * (b % 4) + 2 * (c // 4) - d - (c % 4)) % 7
    f = d + e - 7 * ((a + 11 * d + 22 * e) // 451) + 114
    month = f // 31
    day = f % 31 + 1    
    return dt.date(year, month, day)